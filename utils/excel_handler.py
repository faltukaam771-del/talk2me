"""
excel_handler.py
-----------------
Handles all Excel read/write operations for the chat app using openpyxl.

Excel file structure (chat_data.xlsx):
  Sheet "Rooms":
      room_id | room_name | password_hash | created_at
  Sheet "Messages":
      room_id | username | message | msg_type | file_path | timestamp

Thread-safety: Flask-SocketIO can handle multiple concurrent connections,
so all read/write operations are wrapped in a threading.Lock to avoid
corrupting the .xlsx file when two users send messages at the same time.
"""

import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "chat_data.xlsx")

_lock = threading.Lock()

ROOMS_HEADERS = ["room_id", "room_name", "password_hash", "created_at"]
MESSAGES_HEADERS = ["room_id", "username", "message", "msg_type", "file_path", "timestamp"]


def init_excel():
    """Create the Excel file with proper sheets/headers if it doesn't exist yet."""
    if os.path.exists(EXCEL_FILE):
        return
    wb = Workbook()

    rooms_sheet = wb.active
    rooms_sheet.title = "Rooms"
    rooms_sheet.append(ROOMS_HEADERS)

    messages_sheet = wb.create_sheet("Messages")
    messages_sheet.append(MESSAGES_HEADERS)

    wb.save(EXCEL_FILE)


def _load_wb():
    return load_workbook(EXCEL_FILE)


# ---------------------------------------------------------------------------
# ROOM OPERATIONS
# ---------------------------------------------------------------------------

def create_room(room_id, room_name, password_hash):
    with _lock:
        wb = _load_wb()
        ws = wb["Rooms"]
        ws.append([room_id, room_name, password_hash, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(EXCEL_FILE)


def get_room(room_id):
    """Returns dict with room info, or None if room doesn't exist."""
    with _lock:
        wb = _load_wb()
        ws = wb["Rooms"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == room_id:
                return {
                    "room_id": row[0],
                    "room_name": row[1],
                    "password_hash": row[2],
                    "created_at": row[3],
                }
    return None


def room_exists(room_id):
    return get_room(room_id) is not None


# ---------------------------------------------------------------------------
# MESSAGE OPERATIONS
# ---------------------------------------------------------------------------

def add_message(room_id, username, message, msg_type="text", file_path=""):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with _lock:
        wb = _load_wb()
        ws = wb["Messages"]
        ws.append([room_id, username, message, msg_type, file_path, timestamp])
        wb.save(EXCEL_FILE)
    return timestamp


def get_messages(room_id):
    """Returns list of message dicts for a given room, in chronological order."""
    with _lock:
        wb = _load_wb()
        ws = wb["Messages"]
        messages = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == room_id:
                messages.append({
                    "username": row[1],
                    "message": row[2],
                    "msg_type": row[3],
                    "file_path": row[4],
                    "timestamp": row[5],
                })
    return messages
